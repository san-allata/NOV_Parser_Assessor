#!/usr/bin/env python3
"""Deterministic MRB row builder. Guarantees the exact 149-column A->ES schema
from references/output-schema.md so column order can never drift by hand.

Usage:
    from build_row import build_workbook
    build_workbook(values, "MRB_result.xlsx")

`values` is a flat dict keyed by "Group :: Sub-header" (e.g. "PO :: ERP ID",
"UT Report (Material) :: Status"). Missing keys raise -> forces completeness.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NDT_SUBS = ["PO no.", "ERP ID", "Legacy ID", "QTY", "Test Date", "Inspector Name",
            "Testing Spec", "Inspector qualification", "Expiry Date",
            "Eye Test Examination Date", "Status"]

def _schema():
    cols = []
    def add(g, s): cols.append((g, s))
    add("Case", "Case label"); add("MWI", "MWI")
    for s in ["Supplier Name", "PO line no.", "PO no.", "ERP ID", "Legacy ID", "QTY", "NCL"]:
        add("PO", s)
    for s in ["PO line no.", "PO no.", "ERP ID", "Legacy ID", "QTY", "NCL Level",
              "TN number", "Serial Number", "Status"]:
        add("COC", s)
    add("3.1 MTC", "Status"); add("Balloon Drawing", "Status")
    for s in ["PO no.", "ERP ID", "Legacy ID", "Sampling Inspection QTY",
              "Inspector Name", "Inspection Date", "Status"]:
        add("Dimension Report", s)
    add("Surface Finish", "Status")
    for s in ["PO no.", "ERP ID", "Legacy ID", "QTY"]:
        add("Hardness Report", s)
    for g in ["MPI Report (Material)", "LPI Report (Material)", "UT Report (Material)",
              "RT Report (Material)", "MPI Report (Finished Product)",
              "LPI Report (Finished Product)", "UT Report (Finished Product)",
              "RT Report (Finished Product)", "Ultrasonic Gauge Thickness Report"]:
        for s in NDT_SUBS: add(g, s)
    for s in ["PO no.", "ERP ID", "Legacy ID", "QTY", "Spec", "Status"]:
        add("Coating Report", s)
    for s in ["PO no.", "ERP ID", "Legacy ID", "QTY", "Status"]:
        add("Light Band Report", s)
    add("PWHT Report", "Status")
    add("Concession", "Concession no."); add("Concession", "Status")
    add("NCR", "NCR no."); add("NCR", "Status")
    add("Remarks", "Remarks"); add("Final Result", "Final Result")
    assert len(cols) == 149, f"schema drift: {len(cols)} cols"
    assert get_column_letter(len(cols)) == "ES"
    return cols

def build_workbook(values: dict, out_path: str) -> str:
    cols = _schema()
    missing = [f"{g} :: {s}" for g, s in cols if f"{g} :: {s}" not in values]
    if missing:
        raise KeyError(f"Missing {len(missing)} values (no blank cells allowed): {missing[:8]}...")

    wb = Workbook(); ws = wb.active; ws.title = "MRB Result"
    gf = PatternFill("solid", fgColor="1F4E78"); sf = PatternFill("solid", fgColor="2E75B6")
    wh = Font(color="FFFFFF", bold=True, size=9)
    th = Side(style="thin", color="BFBFBF"); bd = Border(th, th, th, th)
    fills = {"Pass": "C6EFCE", "Fail": "FFC7CE", "Check": "FFEB9C", "Approve": "00B050",
             "Reject": "FF0000", "Yes": "C6EFCE", "No": "F2F2F2", "NA": "F2F2F2"}
    for i, (g, s) in enumerate(cols, 1):
        L = get_column_letter(i)
        for r, val, fl in ((1, g, gf), (2, s, sf)):
            c = ws[f"{L}{r}"]; c.value = val; c.fill = fl; c.font = wh
            c.alignment = Alignment("center", "center", wrap_text=True); c.border = bd
        v = values[f"{g} :: {s}"]
        c3 = ws[f"{L}3"]; c3.value = v; c3.number_format = "@"
        c3.alignment = Alignment("left", "top", wrap_text=True); c3.border = bd
        key = str(v).strip()
        if s in ("Status", "Final Result") and key in fills:
            c3.fill = PatternFill("solid", fgColor=fills[key])
            if key in ("Approve", "Reject"):
                c3.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions[L].width = 44 if s in ("Supplier Name", "Remarks") else 15
    ws.freeze_panes = "C3"
    for r, h in ((1, 30), (2, 30), (3, 90)):
        ws.row_dimensions[r].height = h
    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    print("Schema OK:", len(_schema()), "columns, last =",
          get_column_letter(len(_schema())))

